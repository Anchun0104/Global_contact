import csv
import json
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Professor, Conference, SubConference, Cooperation, ResearchDirection
from auth import get_current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/import_export", tags=["导入导出"])


def _serialize_row(row, fields):
    result = {}
    for f in fields:
        val = getattr(row, f, None)
        if hasattr(val, 'value'):
            val = val.value
        if val is not None:
            result[f] = val
    return result


@router.get("/export/{entity}")
def export_data(
    entity: str,
    format: str = Query("json"),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    if entity == "professors":
        data = db.query(Professor).all()
        fields = ["id", "name", "title", "university", "email", "website", "location", "qs_ranking", "research_direction", "research_keywords", "notes"]
    elif entity == "conferences":
        data = db.query(Conference).all()
        fields = ["id", "name", "year", "location", "field", "description", "status", "total_sessions", "notes"]
    elif entity == "cooperations":
        data = db.query(Cooperation).all()
        fields = ["id", "professor_id", "conference_id", "sub_conference_id", "cooperation_type", "notes"]
    else:
        raise HTTPException(status_code=400, detail="不支持的类型")

    rows = [_serialize_row(d, fields) for d in data]

    if format == "json":
        content = json.dumps(rows, ensure_ascii=False, indent=2, default=str)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={entity}.json"},
        )
    elif format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        content = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={entity}.csv"},
        )
    elif format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = entity
        header_font = Font(bold=True)
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
        for row_idx, r in enumerate(rows, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=r.get(field))
        for col_idx in range(1, len(fields) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={entity}.xlsx"},
        )
    else:
        raise HTTPException(status_code=400, detail="不支持的格式")


@router.post("/import/{entity}")
def import_data(
    entity: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    content = file.file.read()
    filename = file.filename.lower()
    imported = 0
    errors = []

    try:
        if filename.endswith(".json"):
            rows = json.loads(content.decode("utf-8"))
        elif filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith(".xlsx"):
            rows = _parse_xlsx(content)
        else:
            raise HTTPException(status_code=400, detail="不支持的文件格式，请使用 json/csv/xlsx")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

    if entity == "professors":
        for idx, row in enumerate(rows):
            try:
                name = row.get("name") or row.get("姓名")
                if not name:
                    errors.append(f"第 {idx + 1} 行缺少姓名")
                    continue
                p = Professor(
                    name=name,
                    title=row.get("title") or row.get("职称"),
                    university=row.get("university") or row.get("学校"),
                    email=row.get("email") or row.get("邮箱"),
                    website=row.get("website") or row.get("个人主页"),
                    location=row.get("location") or row.get("所在地"),
                    qs_ranking=_int_or_none(row.get("qs_ranking") or row.get("QS排名")),
                    research_direction=row.get("research_direction") or row.get("研究方向"),
                    research_keywords=row.get("research_keywords") or row.get("研究关键词"),
                    notes=row.get("notes") or row.get("备注"),
                )
                db.add(p)
                db.flush()
                direction = p.research_direction
                if direction:
                    existing = db.query(ResearchDirection).filter(ResearchDirection.name == direction).first()
                    if not existing:
                        db.add(ResearchDirection(name=direction))
                imported += 1
            except Exception as e:
                errors.append(f"第 {idx + 1} 行导入失败: {str(e)}")
        db.commit()

    elif entity == "conferences":
        for idx, row in enumerate(rows):
            try:
                name = row.get("name") or row.get("会议名称")
                if not name:
                    errors.append(f"第 {idx + 1} 行缺少会议名称")
                    continue
                c = Conference(
                    name=name,
                    year=_int_or_none(row.get("year") or row.get("年份")),
                    location=row.get("location") or row.get("举办地点"),
                    field=row.get("field") or row.get("领域"),
                    description=row.get("description") or row.get("描述"),
                    total_sessions=_int_or_none(row.get("total_sessions") or row.get("分会场数")),
                    notes=row.get("notes") or row.get("备注"),
                )
                db.add(c)
                imported += 1
            except Exception as e:
                errors.append(f"第 {idx + 1} 行导入失败: {str(e)}")
        db.commit()

    elif entity == "cooperations":
        for idx, row in enumerate(rows):
            try:
                coop = Cooperation(
                    professor_id=_int_or_none(row.get("professor_id") or row.get("教授ID")),
                    conference_id=_int_or_none(row.get("conference_id") or row.get("会议ID")),
                    sub_conference_id=_int_or_none(row.get("sub_conference_id") or row.get("分会场ID")),
                    cooperation_type=row.get("cooperation_type") or row.get("合作类型") or "other",
                    notes=row.get("notes") or row.get("备注"),
                )
                db.add(coop)
                imported += 1
            except Exception as e:
                errors.append(f"第 {idx + 1} 行导入失败: {str(e)}")
        db.commit()
    else:
        raise HTTPException(status_code=400, detail="不支持的数据类型")

    return {"success": True, "message": f"成功导入 {imported} 条记录", "imported_count": imported, "errors": errors}


def _int_or_none(val):
    if val is None:
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _parse_xlsx(content):
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers = [str(h).strip() if h else "" for h in rows_data[0]]
    result = []
    for row in rows_data[1:]:
        item = {}
        for i, val in enumerate(row):
            if i < len(headers):
                item[headers[i]] = str(val).strip() if val is not None else ""
        if any(v for v in item.values()):
            result.append(item)
    return result
